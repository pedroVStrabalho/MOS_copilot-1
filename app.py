"""
MOS Cashflow Copilot - prototype without OpenAI API wired in yet.

What this does
--------------
- Lets you upload PDFs and images OR point to a Dropbox-sync folder
- Creates/updates one Excel workbook with:
    * Base_Transacoes
    * Fluxo_12M
    * Dashboard
    * Config
- Keeps a rolling 12-month structure and auto-creates future months
- Has a simple "chat command" box to tell the system what you want
- Uses a pluggable AI adapter so you can connect OpenAI later
- Includes deterministic parsing rules for common Brazilian docs:
    * NFS-e
    * DARF
    * DAMSP / ISS / TFE
    * generic boletos
- Supports manual review queue for low-confidence items

Run
---
pip install streamlit openpyxl pandas python-dateutil pydantic pillow pymupdf watchdog
streamlit run mos_cashflow_copilot.py

Optional later
--------------
- Add OpenAI API in OpenAIAdapter.extract_transactions()
- Add OCR for scanned images/PDFs if needed
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
from dataclasses import dataclass, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

# =========================
# APP CONFIG
# =========================

APP_NAME = "MOS Cashflow Copilot"
BASE_DIR = Path.cwd()
DATA_DIR = BASE_DIR / "data_mos"
INBOX_DIR = DATA_DIR / "inbox"
PROCESSED_DIR = DATA_DIR / "processed"
EXPORTS_DIR = DATA_DIR / "exports"
DB_DIR = DATA_DIR / "db"
DEFAULT_WORKBOOK = EXPORTS_DIR / "MOS_Cashflow_Master.xlsx"

for folder in [DATA_DIR, INBOX_DIR, PROCESSED_DIR, EXPORTS_DIR, DB_DIR]:
    folder.mkdir(parents=True, exist_ok=True)

MONTHS_AHEAD = 12

# =========================
# DATA MODELS
# =========================

class Transaction(BaseModel):
    source_file: str
    source_hash: str
    extraction_method: str = "rules"
    confidence: float = 0.0

    tipo: str  # Receita / Despesa / Transferencia / Imposto
    categoria: str
    subcategoria: str = ""
    descricao: str

    fornecedor_cliente: str = ""
    documento_numero: str = ""
    competencia: Optional[str] = None  # YYYY-MM
    data_emissao: Optional[str] = None  # YYYY-MM-DD
    vencimento: Optional[str] = None  # YYYY-MM-DD
    data_caixa: Optional[str] = None  # YYYY-MM-DD

    valor: float
    moeda: str = "BRL"

    conta: str = "Conta Corrente"
    centro_custo: str = "Administrativo"
    observacoes: str = ""

    status: str = "Pendente"  # Pendente / Revisado / Lançado
    lancar_no_fluxo: bool = True
    mes_fluxo: Optional[str] = None  # YYYY-MM
    ano_fluxo: Optional[int] = None


class UserCommand(BaseModel):
    raw_text: str
    action: str = "organizar"
    requested_sheet: Optional[str] = None
    filters: Dict[str, Any] = Field(default_factory=dict)
    output_notes: str = ""


class ReviewItem(BaseModel):
    source_file: str
    reason: str
    extracted_data: Dict[str, Any]


# =========================
# HELPERS
# =========================

def brl_to_float(text: str) -> Optional[float]:
    if not text:
        return None
    cleaned = text.strip()
    cleaned = cleaned.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    cleaned = re.sub(r"[^0-9\.-]", "", cleaned)
    if not cleaned:
        return None
    try:
        return float(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return None


def normalize_date_br(text: str) -> Optional[str]:
    if not text:
        return None
    text = text.strip()
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def month_from_date(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    try:
        d = datetime.fromisoformat(date_str).date()
        return f"{d.year:04d}-{d.month:02d}"
    except ValueError:
        return None


def safe_hash_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def month_label(ym: str) -> str:
    d = datetime.strptime(ym + "-01", "%Y-%m-%d")
    nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    return f"{nomes[d.month-1]}/{d.year}"


def ensure_12_month_window(start_month: Optional[str] = None) -> List[str]:
    if start_month:
        start = datetime.strptime(start_month + "-01", "%Y-%m-%d").date().replace(day=1)
    else:
        today = date.today().replace(day=1)
        start = today - relativedelta(months=6)
    return [f"{(start + relativedelta(months=i)).year:04d}-{(start + relativedelta(months=i)).month:02d}" for i in range(MONTHS_AHEAD)]


def classify_command(raw: str) -> UserCommand:
    text = raw.strip().lower()
    action = "organizar"
    requested_sheet = None
    if "dashboard" in text:
        action = "dashboard"
        requested_sheet = "Dashboard"
    elif "fluxo" in text or "caixa" in text:
        action = "fluxo"
        requested_sheet = "Fluxo_12M"
    elif "dre" in text:
        action = "dre"
    elif "excel" in text or "planilha" in text:
        action = "exportar"
    return UserCommand(raw_text=raw, action=action, requested_sheet=requested_sheet, output_notes="interpretação simples")


# =========================
# TEXT EXTRACTION
# =========================

class TextExtractor:
    def extract_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf(path)
        if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            return self._extract_image_stub(path)
        return ""

    def _extract_pdf(self, path: Path) -> str:
        if fitz is None:
            return ""
        texts: List[str] = []
        try:
            doc = fitz.open(path)
            for page in doc:
                texts.append(page.get_text("text"))
            return "\n".join(texts)
        except Exception:
            return ""

    def _extract_image_stub(self, path: Path) -> str:
        # Placeholder. Later you can add OCR here.
        # Example:
        # import pytesseract
        # from PIL import Image
        # return pytesseract.image_to_string(Image.open(path), lang="por")
        return ""


# =========================
# RULE-BASED PARSER
# =========================

class RuleBasedFinanceParser:
    def parse(self, text: str, source_file: str, source_hash: str) -> Tuple[List[Transaction], List[ReviewItem]]:
        text_norm = re.sub(r"[ \t]+", " ", text or "").strip()
        txs: List[Transaction] = []
        review: List[ReviewItem] = []

        if not text_norm:
            review.append(ReviewItem(source_file=source_file, reason="Sem texto extraído; precisa OCR ou revisão manual.", extracted_data={}))
            return txs, review

        lowered = text_norm.lower()

        if "nota fiscal eletrônica de serviços" in lowered or "nfs-e" in lowered:
            tx = self._parse_nfse(text_norm, source_file, source_hash)
            if tx:
                txs.append(tx)
            else:
                review.append(ReviewItem(source_file=source_file, reason="Não consegui estruturar a NFS-e.", extracted_data={"preview": text_norm[:1200]}))
            return txs, review

        if "documento de arrecadação de receitas federais" in lowered or "darf" in lowered:
            tx = self._parse_darf(text_norm, source_file, source_hash)
            if tx:
                txs.append(tx)
            else:
                review.append(ReviewItem(source_file=source_file, reason="Não consegui estruturar o DARF.", extracted_data={"preview": text_norm[:1200]}))
            return txs, review

        if "prefeitura do município de são paulo" in lowered or "secretaria municipal da fazenda" in lowered:
            tx = self._parse_municipal_guide(text_norm, source_file, source_hash)
            if tx:
                txs.append(tx)
            else:
                review.append(ReviewItem(source_file=source_file, reason="Guia municipal não reconhecida totalmente.", extracted_data={"preview": text_norm[:1200]}))
            return txs, review

        if "recibo do pagador" in lowered or "valor do documento" in lowered:
            tx = self._parse_generic_boleto(text_norm, source_file, source_hash)
            if tx:
                txs.append(tx)
            else:
                review.append(ReviewItem(source_file=source_file, reason="Boleto genérico com baixa confiança.", extracted_data={"preview": text_norm[:1200]}))
            return txs, review

        review.append(ReviewItem(source_file=source_file, reason="Documento sem regra pronta. Encaminhar para IA ou revisão manual.", extracted_data={"preview": text_norm[:1200]}))
        return txs, review

    def _find_first(self, pattern: str, text: str, flags: int = re.IGNORECASE) -> Optional[str]:
        m = re.search(pattern, text, flags)
        return m.group(1).strip() if m else None

    def _parse_nfse(self, text: str, source_file: str, source_hash: str) -> Optional[Transaction]:
        numero = self._find_first(r"Número da Nota.*?(\d{3,})", text, re.IGNORECASE | re.DOTALL)
        emissao = self._find_first(r"(\d{2}/\d{2}/\d{4}) \d{2}:\d{2}:\d{2}", text)
        tomador = self._find_first(r"TOMADOR DE SERVIÇOS.*?Nome/Razão Social:\s*(.+?)\s*CPF/CNPJ", text, re.IGNORECASE | re.DOTALL)
        desc = self._find_first(r"DISCRIMINAÇÃO DE SERVIÇOS\s*(.+?)\s*VALOR TOTAL DO SERVIÇO", text, re.IGNORECASE | re.DOTALL)
        valor = self._find_first(r"VALOR TOTAL DO SERVIÇO = R\$ ?([\d\.\,]+)", text)
        iss = self._find_first(r"Valor do ISS \(R\$\)\s*([\d\.\,]+)", text)
        competencia_match = re.search(r"TAXA DE GEST[ÃA]O\s+(\d{2}/\d{4})", text, re.IGNORECASE)
        competencia = None
        if competencia_match:
            mm, yyyy = competencia_match.group(1).split("/")
            competencia = f"{yyyy}-{mm}"

        v = brl_to_float(valor or "")
        iss_v = brl_to_float(iss or "") or 0.0
        if v is None:
            return None

        return Transaction(
            source_file=source_file,
            source_hash=source_hash,
            extraction_method="rules_nfse",
            confidence=0.96,
            tipo="Receita",
            categoria="Receita Operacional",
            subcategoria="Taxa de Gestão",
            descricao=(desc or f"NFS-e {numero or ''}").strip(),
            fornecedor_cliente=tomador or "",
            documento_numero=numero or "",
            competencia=competencia or month_from_date(normalize_date_br(emissao or "")),
            data_emissao=normalize_date_br(emissao or ""),
            vencimento=None,
            data_caixa=None,
            valor=v,
            observacoes=f"ISS destacado aproximado: {iss_v:.2f}",
            status="Pendente",
            mes_fluxo=competencia or month_from_date(normalize_date_br(emissao or "")),
            ano_fluxo=int((competencia or month_from_date(normalize_date_br(emissao or "")) or "2000-01").split("-")[0]),
        )

    def _parse_darf(self, text: str, source_file: str, source_hash: str) -> Optional[Transaction]:
        venc = self._find_first(r"Pagar este documento até\s*(\d{2}/\d{2}/\d{4})", text)
        valor = self._find_first(r"Valor Total do Documento\s*([\d\.\,]+)", text)
        periodo = self._find_first(r"CNPJ Razão Social\s*(\d{2}/\d{2}/\d{4}|[A-Za-zçÇ]+/\d{4})\s+\d{2}/\d{2}/\d{4}", text)
        codigo = self._find_first(r"Composição do Documento de Arrecadação\s*(\d+)\s+([A-ZÇ/ \-\.\w]+?)\s+[\d\.\,]+", text, re.IGNORECASE | re.DOTALL)
        denom_full = None
        m = re.search(r"Composição do Documento de Arrecadação\s*(\d+)\s+(.+?)\s+[\d\.\,]+\s+[\d\.\,]+", text, re.IGNORECASE | re.DOTALL)
        if m:
            codigo = m.group(1)
            denom_full = re.sub(r"\s+", " ", m.group(2)).strip()

        v = brl_to_float(valor or "")
        if v is None:
            return None

        categoria = "Impostos Federais"
        subcategoria = "DARF"
        descricao = denom_full or f"DARF {codigo or ''}".strip()

        tipo = "Imposto"
        comp = None
        if periodo:
            if "/" in periodo and periodo[:2].isdigit():
                comp = datetime.strptime(periodo, "%d/%m/%Y").strftime("%Y-%m")
            else:
                meses = {
                    "janeiro":"01","fevereiro":"02","março":"03","marco":"03","abril":"04","maio":"05","junho":"06",
                    "julho":"07","agosto":"08","setembro":"09","outubro":"10","novembro":"11","dezembro":"12"
                }
                p = periodo.lower()
                for nome, mm in meses.items():
                    if nome in p:
                        ano = re.search(r"(\d{4})", p).group(1)
                        comp = f"{ano}-{mm}"
                        break

        return Transaction(
            source_file=source_file,
            source_hash=source_hash,
            extraction_method="rules_darf",
            confidence=0.95,
            tipo=tipo,
            categoria=categoria,
            subcategoria=subcategoria,
            descricao=descricao,
            fornecedor_cliente="Receita Federal",
            documento_numero=codigo or "",
            competencia=comp,
            data_emissao=None,
            vencimento=normalize_date_br(venc or ""),
            data_caixa=normalize_date_br(venc or ""),
            valor=-v,
            observacoes="DARF importado automaticamente",
            status="Pendente",
            mes_fluxo=month_from_date(normalize_date_br(venc or "")) or comp,
            ano_fluxo=int((month_from_date(normalize_date_br(venc or "")) or comp or "2000-01").split("-")[0]),
        )

    def _parse_municipal_guide(self, text: str, source_file: str, source_hash: str) -> Optional[Transaction]:
        venc = self._find_first(r"Vencimento\s*(\d{2}/\d{2}/\d{4})", text)
        valor = self._find_first(r"Total \(R\$\)\s*([\d\.\,]+)", text)
        receita = self._find_first(r"Receita\s*(.+?)\s*(Outras Informações|Multa|Juros)", text, re.IGNORECASE | re.DOTALL)
        incidencia = self._find_first(r"Incid[êe]ncia\s*([A-Za-z]{3}\s*/\s*\d{4}|\d{2}/\d{4})", text)
        obs = self._find_first(r"Outras Informações\s*(.+?)\s*(Pague|VIA DO CONTRIBUINTE|VIA DO BANCO)", text, re.IGNORECASE | re.DOTALL)

        v = brl_to_float(valor or "")
        if v is None:
            return None

        comp = None
        if incidencia:
            inc = incidencia.replace(" ", "")
            if re.match(r"\d{2}/\d{4}", inc):
                mm, yyyy = inc.split("/")
                comp = f"{yyyy}-{mm}"
            else:
                meses = {"JAN":"01","FEV":"02","MAR":"03","ABR":"04","MAI":"05","JUN":"06","JUL":"07","AGO":"08","SET":"09","OUT":"10","NOV":"11","DEZ":"12"}
                m = re.match(r"([A-Z]{3})/(\d{4})", inc.upper())
                if m:
                    comp = f"{m.group(2)}-{meses[m.group(1)]}"

        rec = (receita or "").strip()
        rec_low = rec.lower()
        if "iss" in rec_low:
            categoria = "Impostos Municipais"
            subcategoria = "ISS"
            descricao = rec
        elif "tfe" in rec_low:
            categoria = "Taxas"
            subcategoria = "TFE"
            descricao = "Taxa de Fiscalização de Estabelecimento"
        else:
            categoria = "Impostos Municipais"
            subcategoria = "Guia Municipal"
            descricao = rec or "Guia Municipal"

        return Transaction(
            source_file=source_file,
            source_hash=source_hash,
            extraction_method="rules_guia_municipal",
            confidence=0.94,
            tipo="Imposto" if subcategoria != "TFE" else "Despesa",
            categoria=categoria,
            subcategoria=subcategoria,
            descricao=descricao,
            fornecedor_cliente="Prefeitura de São Paulo",
            documento_numero="",
            competencia=comp,
            data_emissao=None,
            vencimento=normalize_date_br(venc or ""),
            data_caixa=normalize_date_br(venc or ""),
            valor=-v,
            observacoes=(obs or "").strip(),
            status="Pendente",
            mes_fluxo=month_from_date(normalize_date_br(venc or "")) or comp,
            ano_fluxo=int((month_from_date(normalize_date_br(venc or "")) or comp or "2000-01").split("-")[0]),
        )

    def _parse_generic_boleto(self, text: str, source_file: str, source_hash: str) -> Optional[Transaction]:
        venc = self._find_first(r"Vencimento\s*(\d{2}[\/\.]\d{2}[\/\.]\d{4})", text)
        valor = self._find_first(r"Valor do Documento\s*([\d\.\,]+)", text)
        benef = self._find_first(r"Beneficiário\s*(.+?)\s*(Ag[êe]ncia|Data do Documento|Vencimento)", text, re.IGNORECASE | re.DOTALL)
        detalhe = self._find_first(r"(Detalhamento|Serviço/produto:|Demonstrativo)\s*(.+)", text, re.IGNORECASE)
        if detalhe and isinstance(detalhe, str):
            descricao = detalhe
        else:
            descricao = benef or "Boleto Genérico"

        v = brl_to_float(valor or "")
        if v is None:
            return None

        categoria, subcat = self._infer_boleto_category(text, benef or "", descricao or "")
        return Transaction(
            source_file=source_file,
            source_hash=source_hash,
            extraction_method="rules_boleto",
            confidence=0.78,
            tipo="Despesa",
            categoria=categoria,
            subcategoria=subcat,
            descricao=(descricao or "Boleto").strip(),
            fornecedor_cliente=(benef or "").strip(),
            documento_numero="",
            competencia=None,
            data_emissao=None,
            vencimento=normalize_date_br((venc or "").replace(".", "/")),
            data_caixa=normalize_date_br((venc or "").replace(".", "/")),
            valor=-v,
            observacoes="Boleto genérico; revisar categoria se necessário",
            status="Pendente",
            mes_fluxo=month_from_date(normalize_date_br((venc or "").replace(".", "/"))),
            ano_fluxo=int((month_from_date(normalize_date_br((venc or "").replace(".", "/"))) or "2000-01").split("-")[0]),
        )

    def _infer_boleto_category(self, text: str, beneficiary: str, description: str) -> Tuple[str, str]:
        blob = f"{text} {beneficiary} {description}".lower()
        if "saúde" in blob or "bradesco saúde" in blob or "anbima" in blob:
            return "Pessoal", "Plano de Saúde"
        if "factset" in blob or "agencia estado" in blob or "estado s.a" in blob:
            return "Operacional", "Dados / Research"
        if "xpi" in blob or "xp" in blob:
            return "Operacional", "Plataformas"
        if "vexpenses" in blob:
            return "Operacional", "SaaS"
        if "bhub" in blob:
            return "Administrativo", "Contabilidade"
        return "Operacional", "Fornecedores"


# =========================
# AI ADAPTER (STUB)
# =========================

class BaseAIAdapter:
    def extract_transactions(self, text: str, source_file: str, source_hash: str) -> Tuple[List[Transaction], List[ReviewItem]]:
        raise NotImplementedError


class DummyAIAdapter(BaseAIAdapter):
    def extract_transactions(self, text: str, source_file: str, source_hash: str) -> Tuple[List[Transaction], List[ReviewItem]]:
        return [], [ReviewItem(source_file=source_file, reason="AI adapter não conectado ainda.", extracted_data={})]


class OpenAIAdapter(BaseAIAdapter):
    """
    Placeholder to connect later.

    Suggested future flow:
    1. Send extracted text + user instruction + schema to the model
    2. Ask the model to return JSON matching Transaction[]
    3. Validate with Pydantic
    4. Low-confidence or ambiguous items go to review queue
    """

    def __init__(self, api_key: str, model: str = "gpt-5"):
        self.api_key = api_key
        self.model = model

    def extract_transactions(self, text: str, source_file: str, source_hash: str) -> Tuple[List[Transaction], List[ReviewItem]]:
        # TODO: connect to OpenAI later.
        # Keep this signature to avoid changing the rest of the app.
        return [], [ReviewItem(source_file=source_file, reason="Implementar integração com OpenAI aqui.", extracted_data={"model": self.model})]


# =========================
# STORAGE
# =========================

class JsonStore:
    def __init__(self, base_path: Path):
        self.base_path = base_path
        self.transactions_path = base_path / "transactions.json"
        self.review_path = base_path / "review_queue.json"
        self.commands_path = base_path / "commands.json"
        self._init_files()

    def _init_files(self):
        for p in [self.transactions_path, self.review_path, self.commands_path]:
            if not p.exists():
                p.write_text("[]", encoding="utf-8")

    def load_transactions(self) -> List[Transaction]:
        data = json.loads(self.transactions_path.read_text(encoding="utf-8"))
        return [Transaction(**item) for item in data]

    def save_transactions(self, txs: List[Transaction]) -> None:
        self.transactions_path.write_text(
            json.dumps([t.model_dump() for t in txs], ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

    def load_review(self) -> List[ReviewItem]:
        data = json.loads(self.review_path.read_text(encoding="utf-8"))
        return [ReviewItem(**item) for item in data]

    def save_review(self, items: List[ReviewItem]) -> None:
        self.review_path.write_text(
            json.dumps([i.model_dump() for i in items], ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

    def append_command(self, cmd: UserCommand) -> None:
        data = json.loads(self.commands_path.read_text(encoding="utf-8"))
        data.append({"ts": now_ts(), **cmd.model_dump()})
        self.commands_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# =========================
# EXCEL BUILDER
# =========================

class ExcelCashflowBuilder:
    SHEET_BASE = "Base_Transacoes"
    SHEET_FLUXO = "Fluxo_12M"
    SHEET_DASH = "Dashboard"
    SHEET_CONFIG = "Config"
    SHEET_REVIEW = "Review"

    BASE_HEADERS = [
        "source_file", "source_hash", "extraction_method", "confidence", "tipo",
        "categoria", "subcategoria", "descricao", "fornecedor_cliente",
        "documento_numero", "competencia", "data_emissao", "vencimento",
        "data_caixa", "valor", "moeda", "conta", "centro_custo",
        "observacoes", "status", "lancar_no_fluxo", "mes_fluxo", "ano_fluxo"
    ]

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path

    def build_or_update(self, transactions: List[Transaction], review_items: List[ReviewItem], start_month: Optional[str] = None) -> Path:
        if self.workbook_path.exists():
            wb = load_workbook(self.workbook_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        self._write_base_sheet(wb, transactions)
        self._write_review_sheet(wb, review_items)
        months = ensure_12_month_window(start_month)
        self._write_fluxo_sheet(wb, transactions, months)
        self._write_dashboard_sheet(wb, transactions, months)
        self._write_config_sheet(wb, months)
        wb.save(self.workbook_path)
        return self.workbook_path

    def _write_base_sheet(self, wb, transactions: List[Transaction]):
        if self.SHEET_BASE in wb.sheetnames:
            del wb[self.SHEET_BASE]
        ws = wb.create_sheet(self.SHEET_BASE)

        ws.append(self.BASE_HEADERS)
        for tx in transactions:
            d = tx.model_dump()
            ws.append([d.get(h) for h in self.BASE_HEADERS])

        self._style_header(ws)
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _write_review_sheet(self, wb, review_items: List[ReviewItem]):
        if self.SHEET_REVIEW in wb.sheetnames:
            del wb[self.SHEET_REVIEW]
        ws = wb.create_sheet(self.SHEET_REVIEW)
        headers = ["source_file", "reason", "extracted_data_json"]
        ws.append(headers)
        for item in review_items:
            ws.append([item.source_file, item.reason, json.dumps(item.extracted_data, ensure_ascii=False)])
        self._style_header(ws)
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _write_fluxo_sheet(self, wb, transactions: List[Transaction], months: List[str]):
        if self.SHEET_FLUXO in wb.sheetnames:
            del wb[self.SHEET_FLUXO]
        ws = wb.create_sheet(self.SHEET_FLUXO)

        headers = ["Categoria", "Tipo"] + [month_label(m) for m in months] + ["Total 12M"]
        ws.append(headers)

        categories = [
            ("Receita Operacional", "Receita"),
            ("Impostos Federais", "Saída"),
            ("Impostos Municipais", "Saída"),
            ("Taxas", "Saída"),
            ("Operacional", "Saída"),
            ("Administrativo", "Saída"),
            ("Pessoal", "Saída"),
            ("Transferências", "Movimentação"),
        ]

        month_values = {m: [] for m in months}
        for m in months:
            for tx in transactions:
                tx_month = tx.mes_fluxo or tx.competencia or month_from_date(tx.data_caixa)
                if tx_month == m and tx.lancar_no_fluxo:
                    month_values[m].append(tx)

        for category, tipo_lbl in categories:
            row = [category, tipo_lbl]
            total = 0.0
            for m in months:
                val = sum(tx.valor for tx in month_values[m] if tx.categoria == category)
                row.append(val)
                total += val
            row.append(total)
            ws.append(row)

        ws.append([])
        resumo_rows = [
            ("Receitas", lambda tx: tx.valor if tx.valor > 0 else 0),
            ("Saídas", lambda tx: tx.valor if tx.valor < 0 else 0),
            ("Saldo Líquido", lambda tx: tx.valor),
        ]
        for label, func in resumo_rows:
            row = [label, "Resumo"]
            total = 0.0
            for m in months:
                val = sum(func(tx) for tx in month_values[m] if tx.lancar_no_fluxo)
                row.append(val)
                total += val
            row.append(total)
            ws.append(row)

        self._style_header(ws)
        self._format_currency_cols(ws, start_col=3, end_col=2 + len(months) + 1)
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _write_dashboard_sheet(self, wb, transactions: List[Transaction], months: List[str]):
        if self.SHEET_DASH in wb.sheetnames:
            del wb[self.SHEET_DASH]
        ws = wb.create_sheet(self.SHEET_DASH)

        receitas = sum(tx.valor for tx in transactions if tx.valor > 0)
        saidas = sum(abs(tx.valor) for tx in transactions if tx.valor < 0)
        impostos = sum(abs(tx.valor) for tx in transactions if "Impostos" in tx.categoria or tx.subcategoria in {"ISS", "TFE"})
        saldo = receitas - saidas
        carga = (impostos / receitas) if receitas else 0
        margem = (saldo / receitas) if receitas else 0

        ws["A1"] = APP_NAME
        ws["A2"] = "Atualizado em"
        ws["B2"] = now_ts()

        cards = [
            ("Receitas", receitas),
            ("Saídas", -saidas),
            ("Saldo Líquido", saldo),
            ("Impostos", -impostos),
            ("Carga Tributária", carga),
            ("Margem Líquida", margem),
        ]
        row = 4
        for title, value in cards:
            ws[f"A{row}"] = title
            ws[f"B{row}"] = value
            row += 1

        row += 1
        ws[f"A{row}"] = "Resumo por Mês"
        row += 1
        ws.append(["Mês", "Receitas", "Saídas", "Saldo"])
        start_summary = row + 1

        for m in months:
            rec = sum(tx.valor for tx in transactions if (tx.mes_fluxo or tx.competencia or month_from_date(tx.data_caixa)) == m and tx.valor > 0)
            sai = sum(abs(tx.valor) for tx in transactions if (tx.mes_fluxo or tx.competencia or month_from_date(tx.data_caixa)) == m and tx.valor < 0)
            ws.append([month_label(m), rec, sai, rec - sai])

        self._style_header(ws, row=row)
        self._format_currency_cols(ws, start_col=2, end_col=4, start_row=4, end_row=9)
        self._format_percent_cell(ws["B8"])
        self._format_percent_cell(ws["B9"])
        self._format_currency_cols(ws, start_col=2, end_col=4, start_row=start_summary, end_row=start_summary + len(months) - 1)
        self._auto_width(ws)

    def _write_config_sheet(self, wb, months: List[str]):
        if self.SHEET_CONFIG in wb.sheetnames:
            del wb[self.SHEET_CONFIG]
        ws = wb.create_sheet(self.SHEET_CONFIG)
        ws.append(["Parâmetro", "Valor"])
        ws.append(["Janela_12M_inicio", months[0]])
        ws.append(["Janela_12M_fim", months[-1]])
        ws.append(["Conta_corrente_minima", 1000])
        ws.append(["Observação", "Você pode mudar a janela de meses e reconstruir a planilha."])
        self._style_header(ws)
        self._auto_width(ws)

    def _style_header(self, ws, row: int = 1):
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[row]:
            cell.fill = fill
            cell.font = font

    def _format_currency_cols(self, ws, start_col: int, end_col: int, start_row: int = 2, end_row: Optional[int] = None):
        if end_row is None:
            end_row = ws.max_row
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(r, c).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

    def _format_percent_cell(self, cell):
        cell.number_format = '0.00%'

    def _auto_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(val))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)


# =========================
# PIPELINE
# =========================

class DocumentIngestionService:
    def __init__(self, extractor: TextExtractor, parser: RuleBasedFinanceParser, ai_adapter: BaseAIAdapter, store: JsonStore):
        self.extractor = extractor
        self.parser = parser
        self.ai_adapter = ai_adapter
        self.store = store

    def ingest_file(self, path: Path, use_ai_fallback: bool = False) -> Tuple[List[Transaction], List[ReviewItem]]:
        raw = path.read_bytes()
        source_hash = safe_hash_bytes(raw)
        existing = self.store.load_transactions()
        if any(tx.source_hash == source_hash for tx in existing):
            return [], [ReviewItem(source_file=path.name, reason="Arquivo já importado; ignorado para evitar duplicidade.", extracted_data={"hash": source_hash})]

        text = self.extractor.extract_text(path)
        txs, review = self.parser.parse(text, path.name, source_hash)

        if use_ai_fallback and not txs:
            ai_txs, ai_review = self.ai_adapter.extract_transactions(text, path.name, source_hash)
            txs.extend(ai_txs)
            review.extend(ai_review)

        return txs, review

    def persist_results(self, txs: List[Transaction], review: List[ReviewItem]) -> None:
        current_txs = self.store.load_transactions()
        current_review = self.store.load_review()
        current_txs.extend(txs)
        current_review.extend(review)
        self.store.save_transactions(current_txs)
        self.store.save_review(current_review)


# =========================
# DROPBOX / FOLDER WATCH
# =========================

def scan_dropbox_folder(folder: Path) -> List[Path]:
    if not folder.exists():
        return []
    files = []
    for p in folder.glob("*"):
        if p.is_file() and p.suffix.lower() in {".pdf", ".png", ".jpg", ".jpeg", ".webp"}:
            files.append(p)
    return sorted(files)


# =========================
# STREAMLIT UI
# =========================

def init_session():
    if "dropbox_dir" not in st.session_state:
        st.session_state.dropbox_dir = str(INBOX_DIR)
    if "use_ai_fallback" not in st.session_state:
        st.session_state.use_ai_fallback = False
    if "window_start_month" not in st.session_state:
        today = date.today().replace(day=1)
        st.session_state.window_start_month = (today - relativedelta(months=6)).strftime("%Y-%m")


def render_sidebar(store: JsonStore):
    st.sidebar.header("Configuração")
    st.session_state.dropbox_dir = st.sidebar.text_input("Pasta Dropbox / Inbox", value=st.session_state.dropbox_dir)
    st.session_state.use_ai_fallback = st.sidebar.checkbox("Usar fallback de IA quando regra falhar", value=st.session_state.use_ai_fallback)
    st.session_state.window_start_month = st.sidebar.text_input("Mês inicial da janela 12M (YYYY-MM)", value=st.session_state.window_start_month)
    st.sidebar.caption("Você pode apontar essa pasta para uma pasta local sincronizada pelo Dropbox.")

    txs = store.load_transactions()
    review = store.load_review()
    st.sidebar.metric("Transações salvas", len(txs))
    st.sidebar.metric("Itens para revisão", len(review))
    st.sidebar.metric("Workbook", "Pronto" if DEFAULT_WORKBOOK.exists() else "Ainda não")

def render_top():
    st.title(APP_NAME)
    st.write("Faça upload de PDFs/fotos, organize tudo em um Excel único e mantenha um fluxo de caixa inteligente de 12 meses.")

def save_uploaded_files(uploaded_files: List[Any]) -> List[Path]:
    saved = []
    for up in uploaded_files:
        target = INBOX_DIR / up.name
        target.write_bytes(up.getbuffer())
        saved.append(target)
    return saved

def run_ingestion(paths: List[Path], service: DocumentIngestionService) -> Tuple[int, int]:
    added = 0
    review_count = 0
    for p in paths:
        txs, review = service.ingest_file(p, use_ai_fallback=st.session_state.use_ai_fallback)
        service.persist_results(txs, review)
        added += len(txs)
        review_count += len(review)
        try:
            shutil.copy2(p, PROCESSED_DIR / p.name)
        except Exception:
            pass
    return added, review_count

def rebuild_workbook(store: JsonStore):
    builder = ExcelCashflowBuilder(DEFAULT_WORKBOOK)
    path = builder.build_or_update(
        transactions=store.load_transactions(),
        review_items=store.load_review(),
        start_month=st.session_state.window_start_month
    )
    return path

def chat_command_area(store: JsonStore):
    st.subheader("Comando em linguagem natural")
    cmd_text = st.text_area(
        "Diga o que você quer",
        value="Organize tudo no fluxo anual, separe impostos, atualize o dashboard e deixe 12 meses prontos."
    )
    if st.button("Interpretar comando"):
        cmd = classify_command(cmd_text)
        store.append_command(cmd)
        st.success(f"Ação entendida: {cmd.action}")
        st.json(cmd.model_dump(), expanded=False)

def uploads_area(store: JsonStore, service: DocumentIngestionService):
    st.subheader("1) Upload de PDFs e fotos")
    uploaded = st.file_uploader("Arraste PDFs/imagens", type=["pdf", "png", "jpg", "jpeg", "webp"], accept_multiple_files=True)
    if st.button("Processar uploads"):
        if not uploaded:
            st.warning("Envie pelo menos um arquivo.")
        else:
            paths = save_uploaded_files(uploaded)
            added, review_count = run_ingestion(paths, service)
            wb = rebuild_workbook(store)
            st.success(f"Processados. Novas transações: {added}. Itens em revisão: {review_count}.")
            st.download_button("Baixar Excel atualizado", data=wb.read_bytes(), file_name=wb.name)

def dropbox_area(store: JsonStore, service: DocumentIngestionService):
    st.subheader("2) Pasta Dropbox / sincronização local")
    folder = Path(st.session_state.dropbox_dir)
    st.caption(f"Pasta monitorada: {folder}")
    files = scan_dropbox_folder(folder)
    st.write(f"Arquivos encontrados: {len(files)}")
    if files:
        st.dataframe(pd.DataFrame({"arquivo": [f.name for f in files]}), use_container_width=True)
    if st.button("Importar arquivos da pasta"):
        added, review_count = run_ingestion(files, service)
        wb = rebuild_workbook(store)
        st.success(f"Importação concluída. Novas transações: {added}. Itens em revisão: {review_count}.")
        st.download_button("Baixar Excel atualizado", data=wb.read_bytes(), file_name=wb.name)

def data_area(store: JsonStore):
    st.subheader("3) Base de transações")
    txs = store.load_transactions()
    if txs:
        df = pd.DataFrame([t.model_dump() for t in txs])
        st.dataframe(df, use_container_width=True)
    else:
        st.info("Ainda não há transações salvas.")

    st.subheader("4) Itens para revisão")
    items = store.load_review()
    if items:
        rdf = pd.DataFrame([i.model_dump() for i in items])
        st.dataframe(rdf, use_container_width=True)
    else:
        st.success("Sem pendências de revisão.")

def workbook_area(store: JsonStore):
    st.subheader("5) Excel mestre")
    if st.button("Reconstruir workbook agora"):
        wb = rebuild_workbook(store)
        st.success("Workbook reconstruído.")
        st.download_button("Baixar Excel mestre", data=wb.read_bytes(), file_name=wb.name)

    if DEFAULT_WORKBOOK.exists():
        st.download_button("Baixar último Excel gerado", data=DEFAULT_WORKBOOK.read_bytes(), file_name=DEFAULT_WORKBOOK.name)

def manual_entry_area(store: JsonStore):
    st.subheader("6) Lançamento manual rápido")
    with st.form("manual_tx"):
        col1, col2 = st.columns(2)
        tipo = col1.selectbox("Tipo", ["Receita", "Despesa", "Imposto", "Transferencia"])
        categoria = col2.text_input("Categoria", value="Operacional")
        descricao = st.text_input("Descrição")
        fornecedor_cliente = st.text_input("Fornecedor / Cliente")
        competencia = st.text_input("Competência (YYYY-MM)", value=date.today().strftime("%Y-%m"))
        data_caixa = st.text_input("Data caixa (YYYY-MM-DD)", value=date.today().isoformat())
        valor = st.number_input("Valor", value=0.0, step=100.0)
        submitted = st.form_submit_button("Adicionar lançamento")
        if submitted:
            sign_value = valor if tipo == "Receita" else -abs(valor)
            tx = Transaction(
                source_file="manual_entry",
                source_hash=f"manual_{hash((descricao, data_caixa, valor, now_ts()))}",
                extraction_method="manual",
                confidence=1.0,
                tipo=tipo,
                categoria=categoria,
                descricao=descricao,
                fornecedor_cliente=fornecedor_cliente,
                competencia=competencia,
                data_caixa=data_caixa,
                valor=sign_value,
                mes_fluxo=competencia or month_from_date(data_caixa),
                ano_fluxo=int((competencia or month_from_date(data_caixa) or "2000-01").split("-")[0]),
                status="Revisado",
            )
            txs = store.load_transactions()
            txs.append(tx)
            store.save_transactions(txs)
            rebuild_workbook(store)
            st.success("Lançamento adicionado e workbook atualizado.")

def main():
    st.set_page_config(page_title=APP_NAME, layout="wide")
    init_session()

    store = JsonStore(DB_DIR)
    extractor = TextExtractor()
    parser = RuleBasedFinanceParser()
    ai_adapter = DummyAIAdapter()
    service = DocumentIngestionService(extractor, parser, ai_adapter, store)

    render_sidebar(store)
    render_top()
    chat_command_area(store)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Uploads",
        "Dropbox / Pasta",
        "Base & Revisão",
        "Workbook",
        "Manual"
    ])

    with tab1:
        uploads_area(store, service)
    with tab2:
        dropbox_area(store, service)
    with tab3:
        data_area(store)
    with tab4:
        workbook_area(store)
    with tab5:
        manual_entry_area(store)

    st.divider()
    st.caption(
        "Próximos upgrades: OCR para imagens/PDFs escaneados, integração OpenAI, "
        "regras avançadas por fornecedor, conciliação bancária e chatbot com memória."
    )

if __name__ == "__main__":
    main()
